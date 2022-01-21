# -*- coding: utf-8 -*-
"""
Created on Fri Jan 21 10:11:54 2022

@author: okmen
"""


"""
# selenium ile olan kısım
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import selenium
from  selenium import webdriver
from selenium.webdriver.common.by import By
options = Options()
options.binary_location = "C:\Program Files\Google\Chrome\Application\chrome.exe"
driver = webdriver.Chrome(chrome_options = options, executable_path=r'chromedriver.exe')
driver.get('https://www.defacto.com.tr/regular-fit-bisiklet-yaka-basic-pamuklu-sweatshirt-2409516')
print("Chrome Browser Invoked successfully")
driver.find_element("/html/body/div[1]/div[3]/form/div[1]/div[1]/div[1]/div/div[2]/input").click()
elem = driver.find_element_by_xpath('/html/body/div[1]/div[3]/form/div[1]/div[1]/div[3]/center/input[1]')
elem.click()
elem = driver.find_element_by_xpath('/html/body/div[1]/div[3]/form/div[1]/div[1]/div[3]/center/input[1]')
elem.click()

#selenium olmadan

"""

from bs4 import BeautifulSoup

import urllib.request as urllib2

import openpyxl

import pandas as pd

with open("Linkler.txt","r") as dosya:
        Linkler = dosya.readlines()   


linkler=[]

for ss in range(0,len(Linkler)):
    linkler.append(Linkler[ss].split("\n")[0])
    pass


 
#excel için  açma
    
    
kitap = openpyxl.load_workbook("haydicanım.xlsx")
sayfa = kitap.get_sheet_by_name("Sheet1")
    

global sira
sira=2
for link_geliyo in linkler:
        
    c= urllib2.urlopen(link_geliyo)
    contents=c.read()
    soup=BeautifulSoup(contents)
    soup=str(soup)


   
    
    
    
    
    #ANA ÜRÜN
    ########################################################################
    #isim almak için
    
    sayac=0
    ana_bilesen=soup.split("title")[2]
    ana_bilesen=ana_bilesen[2:]
    for i in ana_bilesen :
        if i == " ":
            sayac=sayac+1
        else:
            break
    ana_bilesen=ana_bilesen[sayac:]
    ana_bilesen=ana_bilesen.split(" ")
    ana_urun_kodu_yeri=ana_bilesen.index("|")
    
    #ana adi olusturma / birleştirme
    
    ana_adi=""
    for k in range(ana_urun_kodu_yeri-1):
        ana_adi=ana_adi+" "+ana_bilesen[k]
    #bosluk silme
    if ana_adi[0]==" ":
        ana_adi=ana_adi[1:]
        
    #####################################################################33
    
    
    ############################################################################33
    
    #fiyat belirlemek için alan
    
    try:
            
        fiyat_bilesen=soup.split("normalPrice")[1]
        fiyat_bilesen=fiyat_bilesen[2:]
        fiyat_bilesen=fiyat_bilesen.split("₺")[1].split("</span>")[0]
        fiyat_i=fiyat_bilesen
        fiyat_n=fiyat_bilesen
    except:
        fiyat_bilesen=soup.split("insteadPrice")[1]
        fiyat_bilesen=fiyat_bilesen[2:]
        fiyat_bilesen_normal=fiyat_bilesen.split("₺")[1].split("<")[0]
        fiyat_n=fiyat_bilesen_normal
        ####################################################################
        fiyat_bilesen=soup.split("newPrice")[1]
        fiyat_bilesen=fiyat_bilesen[2:]
        fiyat_bilesen_normal=fiyat_bilesen.split("₺")[1].split("<")[0]
        fiyat_i=fiyat_bilesen_normal
        
        
        pass
    
    
    ################################################################33
    
    ################################################################33
    #resim linkleri için
    #ana_adi=ana_adi+" "+ana_bilesen[k]
    resim_bilesen=soup.split("alt-src")
    resim_bilesen.pop(0)
    resim_bilesen_yazma=""
    for j in range(len(resim_bilesen)):
       
        resim_bilesen_yazma=resim_bilesen_yazma + "," + resim_bilesen[j][2:].split('"')[0]
    if resim_bilesen_yazma[0]==",":
            resim_bilesen_yazma=resim_bilesen_yazma[1:]
    
    
    ################################################################33
    
    ################################################################33
    #kategori bulma
    kategori_bilesen=soup.split("category")
    kategori_bilesen=kategori_bilesen[1].split("'")[2]
    
    ################################################################33
    
    
    ################################################################33
    ##### açıklama
    try:
        
        aciklama_bilesen=soup.split("alt-text")
        a11=aciklama_bilesen[1].split("</p>")
        a1=a11[0][2:]
        try:
                
            a1_x=a1[0].split(">")[1]
        except:
            a1_x=a1
        try:
            
            a2=a11[1].split(">")[1]
        except:
            a2=a11
            
        a3=aciklama_bilesen[2].split("</p>")[0][2:]
        kisa_aciklama=a1_x
        uzun_aciklama=a1_x+" "+a2+a3
        try:
            a4=aciklama_bilesen[3].split("</p>")[0][2:]
            uzun_aciklama=a1_x+" "+a2+a3+a4
        except:
            pass
    except Exception as e:
        print("hata1:",e)
    
    
    
    
    ################################################################33
    
    
    
    
    
    
    
    
    
    
    
    
    #yazdırılacaklar
    
    ana_adi=ana_adi
    ana_urunkodu=ana_bilesen[ana_urun_kodu_yeri-1]
    fiyat_i=fiyat_i
    fiyat_n=fiyat_n
    resim=resim_bilesen_yazma
    kategori=kategori_bilesen
    
    
    
    
    
    
    try:
            
        sayfa.cell(row=sira,column=9,value=kisa_aciklama)
        sayfa.cell(row=sira,column=10,value=uzun_aciklama)
    except Exception as e:
        print("hata 2:",e)
    
    sayfa.cell(row=sira,column=4,value=ana_urunkodu)
    sayfa.cell(row=sira,column=5,value=ana_adi)
    sayfa.cell(row=sira,column=26,value=fiyat_i)
    sayfa.cell(row=sira,column=27,value=fiyat_n)
    sayfa.cell(row=sira,column=28,value=kategori)
    sayfa.cell(row=sira,column=31,value=resim)
    sayfa.cell(row=sira,column=7,value="0")    
    sayfa.cell(row=sira,column=6,value="1")
    sayfa.cell(row=sira,column=3,value="variable")
    sayfa.cell(row=sira,column=8,value="visible")
    sayfa.cell(row=sira,column=13,value="taxable")
    sayfa.cell(row=sira,column=15,value="1")
    sayfa.cell(row=sira,column=19,value="0")
    sayfa.cell(row=sira,column=18,value="0")
    sayfa.cell(row=sira,column=24,value="1")
    sayfa.cell(row=sira,column=41,value="size")
    sayfa.cell(row=sira,column=43,value="1")
    sayfa.cell(row=sira,column=44,value="1")
    
    #varyasyon
    
    
    NeKadarKalmis=soup.split("size-items")
    NeKadarKalmis=NeKadarKalmis[1].split("<li>")
    NeKadarKalmis.pop(0)
    
    #varyasyonları ayıklama
    sayac_varyasyon_zamani=0
    sira=sira+1
    hafiza_beden=""
    hafiza_beden_icin_sira=sira-1
    while 1:
        try: 
               
            bulma_kismi=NeKadarKalmis[sayac_varyasyon_zamani].split(" ")
            sku=bulma_kismi[1][6:-1]
            stok_adet=bulma_kismi[4][12:-1]
            
            fiyat_i=fiyat_i
            fiyat_n=fiyat_n
            ana_urunkodu
            beden=bulma_kismi[5].split(">")[1].split("<")[0]
            adi=ana_adi+"- "+beden
            
            hafiza_beden=hafiza_beden+","+beden
            #yazdırmaya başladımmmmm
            
            sayfa.cell(row=sira,column=44,value="1")
            sayfa.cell(row=sira,column=41,value="size")
            sayfa.cell(row=sira,column=24,value="0")
            sayfa.cell(row=sira,column=19,value="0")
            sayfa.cell(row=sira,column=18,value="0")
            sayfa.cell(row=sira,column=15,value="1")
            sayfa.cell(row=sira,column=14,value="parent")
            sayfa.cell(row=sira,column=13,value="taxable")
            sayfa.cell(row=sira,column=8,value="visible")
            sayfa.cell(row=sira,column=7,value="0") 
            sayfa.cell(row=sira,column=6,value="1")
            sayfa.cell(row=sira,column=3,value="variation")
            
            
            
            
            
            
            
            
            
            
            
            
            sayfa.cell(row=sira,column=42,value=beden)
            sayfa.cell(row=sira,column=4,value=sku)
            sayfa.cell(row=sira,column=5,value=adi)
            sayfa.cell(row=sira,column=16,value=stok_adet)
            sayfa.cell(row=sira,column=26,value=fiyat_i)
            sayfa.cell(row=sira,column=27,value=fiyat_n)
            sayfa.cell(row=sira,column=34,value=ana_urunkodu)
            sira=sira+1
            sayac_varyasyon_zamani=sayac_varyasyon_zamani+1
            
        except:
            if hafiza_beden[0]==",":
                hafiza_beden=hafiza_beden[1:]
            sayfa.cell(row=hafiza_beden_icin_sira,column=42,value=hafiza_beden)
            break
    
kitap.save("haydicanım.xlsx")
kitap.close()
    
df = pd.read_excel('haydicanım.xlsx')
df.to_csv("haydicanım.csv", sep=',')
    
    
    
    #yazdırılacaklar
"""
    sku
    stok_adet
    ana_urunkodu
    beden
    adi
    fiyat_i
    fiyat_n
"""
    
    
    
    
    














"""


koton bot


"""
























































